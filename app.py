import streamlit as st
from openpyxl import load_workbook

# Display a file uploader widget
xlsx_file = st.file_uploader("Upload XLSX file", type=["xlsx"])

if xlsx_file is not None:
    # Load the XLSX file
    wb = load_workbook(xlsx_file)

    # Select the first sheet
    sheet = wb.active

    # Read the data from the sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Display the data in Streamlit
    st.write(data)
